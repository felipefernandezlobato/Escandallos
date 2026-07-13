"""
parse_excel.py — Parse COCINA - BRÜ Excel file ("ESTUDIO ingredientes " sheet)

Extracts:
  - Ingredient rows: short name, full name, category, supplier prices
  - Date columns (I onwards): header dates
  - Order matrix cells: value + any comment (stock note)
  - Summary counts

Output: JSON saved to excel_data.json in same directory
"""

import json
import re
import sys
import warnings
from datetime import date, datetime
from pathlib import Path

warnings.filterwarnings("ignore")  # suppress openpyxl pivot cache warnings

import openpyxl
from openpyxl.utils import get_column_letter

# ── Paths ────────────────────────────────────────────────────────────────────
EXCEL_PATH = Path("/Users/fernaf41/projects/Tests/Escandallos/COCINA - BRÜ  (3).xlsx")
OUTPUT_PATH = Path(__file__).parent / "excel_data.json"
SHEET_NAME = "ESTUDIO ingredientes "  # note trailing space

# ── Date parsing ─────────────────────────────────────────────────────────────
# Dates in row 2 appear in several formats:
#   float  → 9.7  means 9 July 2025  (DD.MM, assumed current year context)
#   str    → "20.03'" or "13.03'" etc  (DD.MM' with trailing apostrophe, year 2025)
#   Older dates (Oct-Dec) → likely 2024
#   We'll infer year from column order: rightmost = oldest

def _parse_date_header(raw):
    """Convert a raw header value to ISO date string (YYYY-MM-DD)."""
    if raw is None:
        return None

    # Handle datetime objects directly (e.g. 2026-07-09 00:00:00)
    if isinstance(raw, datetime):
        return raw.date().isoformat()

    s = str(raw).strip().rstrip("'")

    # Try float-style "DD.MM" e.g. 9.7 → "9.7"
    # These come from numbers like 9.7 (9th July), 2.7, 25.6 ...
    parts = s.split(".")
    if len(parts) == 2:
        try:
            day = int(parts[0])
            month = int(parts[1])
            if 1 <= day <= 31 and 1 <= month <= 12:
                # Assign year: months 10-12 = 2025, months 1-9 = 2026
                year = 2025 if month >= 10 else 2026
                return date(year, month, day).isoformat()
        except (ValueError, TypeError):
            pass

    return None  # couldn't parse


def _clean_comment(raw_text: str) -> str:
    """Strip the Excel threaded-comment boilerplate, return just the note."""
    if not raw_text:
        return ""
    # Threaded comment format has a header block ending with "Comment:\n\t"
    marker = "Comment:\n\t"
    idx = raw_text.find(marker)
    if idx != -1:
        return raw_text[idx + len(marker):].strip()
    return raw_text.strip()


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    print(f"Loading {EXCEL_PATH.name} …")
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")

    ws = wb[SHEET_NAME]
    print(f"Sheet dimensions: {ws.dimensions}  (max_row={ws.max_row}, max_col={ws.max_column})")

    # ── 1. Parse header row 2 to find date columns ──────────────────────────
    # Columns A=1 B=2 C=3 D=4 E=5 F=6 G=7 H=8 → I=9 onwards are date columns
    SUPPLIER_COLS = {
        "saviva":   5,   # E
        "caporaso": 6,   # F
        "pfaff":    7,   # G
        "prodega":  8,   # H
    }
    DATE_COL_START = 9   # column I

    date_cols: list[dict] = []   # [{col_index, date_str, col_letter}]
    for col_idx in range(DATE_COL_START, ws.max_column + 1):
        raw = ws.cell(row=2, column=col_idx).value
        date_str = _parse_date_header(raw)
        if date_str:
            date_cols.append({
                "col_index": col_idx,
                "col_letter": get_column_letter(col_idx),
                "date": date_str,
                "raw_header": str(raw),
            })

    # Sort chronologically (rightmost = oldest, leftmost = newest)
    date_cols.sort(key=lambda d: d["date"])

    print(f"\nDate columns found: {len(date_cols)}")
    print(f"  Earliest: {date_cols[0]['date']}  (col {date_cols[0]['col_letter']})")
    print(f"  Latest:   {date_cols[-1]['date']} (col {date_cols[-1]['col_letter']})")

    # ── 2. Parse ingredient rows (row 3 onwards) ────────────────────────────
    ingredients: list[dict] = []
    orders: list[dict] = []

    orders_filled = 0
    comments_found = 0

    last_category = None   # carry forward C if current row has none

    for row_idx in range(3, ws.max_row + 1):
        short_name = ws.cell(row=row_idx, column=1).value
        full_name  = ws.cell(row=row_idx, column=2).value
        category   = ws.cell(row=row_idx, column=3).value
        units      = ws.cell(row=row_idx, column=4).value

        # Skip completely empty rows
        if short_name is None and full_name is None:
            continue

        # Normalise strings
        short_name = str(short_name).strip() if short_name else None
        full_name  = str(full_name).strip()  if full_name  else None
        category   = str(category).strip()   if category   else last_category
        units      = str(units).strip()      if units      else None

        last_category = category

        # Supplier prices (E-H)
        prices = {}
        for sup_name, col_idx in SUPPLIER_COLS.items():
            cell = ws.cell(row=row_idx, column=col_idx)
            val  = cell.value
            if val is not None:
                try:
                    prices[sup_name] = float(val)
                except (ValueError, TypeError):
                    prices[sup_name] = str(val)
            else:
                prices[sup_name] = None

        ingredient = {
            "row":        row_idx,
            "short_name": short_name,
            "full_name":  full_name,
            "category":   category,
            "units":      units,
            "prices":     prices,
        }
        ingredients.append(ingredient)

        # Order cells for this ingredient across all date columns
        for dc in date_cols:
            cell = ws.cell(row=row_idx, column=dc["col_index"])
            val  = cell.value
            comment_raw  = cell.comment.text if cell.comment else None
            comment_clean = _clean_comment(comment_raw) if comment_raw else None

            if val is None and comment_clean is None:
                continue   # truly empty, skip

            orders_filled += 1
            if comment_clean:
                comments_found += 1

            # Normalise order value
            if val is not None:
                val_str = str(val).strip() if not isinstance(val, (int, float)) else val
            else:
                val_str = None

            orders.append({
                "row":      row_idx,
                "date":     dc["date"],
                "col":      dc["col_letter"],
                "value":    val_str,
                "comment":  comment_clean,
            })

    print(f"\nIngredients parsed:  {len(ingredients)}")
    print(f"Orders (non-empty):  {orders_filled}")
    print(f"  of which with comment: {comments_found}")

    # ── 3. Build output structure ────────────────────────────────────────────
    output = {
        "ingredients": ingredients,
        "dates": [dc["date"] for dc in date_cols],
        "date_columns": date_cols,
        "orders": orders,
        "summary": {
            "ingredients": len(ingredients),
            "dates": len(date_cols),
            "orders_filled": orders_filled,
            "comments": comments_found,
        },
    }

    # ── 4. Save JSON ─────────────────────────────────────────────────────────
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2, default=str)

    print(f"\nJSON saved → {OUTPUT_PATH}")

    # ── 5. Spot-check printout ───────────────────────────────────────────────
    print("\n=== SAMPLE INGREDIENTS (first 5) ===")
    for ing in ingredients[:5]:
        print(f"  Row {ing['row']:>3}: [{ing['category']}] "
              f"{ing['short_name'] or '—':15s} | {ing['full_name']}")
        p = ing['prices']
        price_str = ", ".join(
            f"{k}={v}" for k, v in p.items() if v is not None
        )
        if price_str:
            print(f"           prices: {price_str}")

    print("\n=== SAMPLE ORDERS WITH COMMENTS (first 10) ===")
    orders_with_comments = [o for o in orders if o["comment"]]
    for o in orders_with_comments[:10]:
        # Find ingredient name
        ing_name = next(
            (i["full_name"] for i in ingredients if i["row"] == o["row"]),
            f"row {o['row']}"
        )
        print(f"  {o['date']} | {ing_name[:40]:40s} | val={repr(o['value'])} | note={repr(o['comment'])}")

    print("\n=== SUMMARY ===")
    for k, v in output["summary"].items():
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
